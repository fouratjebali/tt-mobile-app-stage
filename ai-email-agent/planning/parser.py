from __future__ import annotations

import hashlib
import re
import unicodedata
from collections.abc import Iterable
from datetime import date, datetime, time
from io import BytesIO
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from planning.models import PlanningFileResult, PlanningParticipant, TrainingSession


HEADER_ALIASES: dict[str, tuple[str, ...]] = {
    "code_session": ("code session",),
    "lms_session_number": ("n session lms", "no session lms", "numero session lms"),
    "malek_number": ("n malek", "no malek", "numero malek"),
    "status": ("etat",),
    "axis": ("axe strategique de la formation",),
    "domain": ("domaine d'activite", "domaine d activite"),
    "project": ("projet",),
    "training_type": ("type",),
    "training_mode": ("mode formation",),
    "certification_nature": ("nature formation",),
    "module_code": ("code module",),
    "module": ("module",),
    "cabinet": ("cabinet",),
    "trainer": ("formateur", "formateur(s) retenu(s)", "formateurs retenus"),
    "selected_trainer": ("formateur designe",),
    "year": ("annee",),
    "month": ("mois",),
    "week": ("semaine",),
    "duration_days": ("duree (j)", "duree"),
    "start_date": ("date debut",),
    "end_date": ("date fin",),
    "schedule": ("horaire",),
    "hours_per_day": ("nbre d'heures/par jour", "nombre d'heures par jour"),
    "total_hours": ("total nb heures",),
    "location": ("lieu de formation",),
    "accommodation_location": ("lieu hebergement",),
    "responsible_engagement": ("responsable engagement",),
    "candidate_count": ("nbre candidats", "nombre candidats"),
    "matricule": ("matricules", "matricule"),
    "full_name": ("nom & prenom", "nom et prenom", "nom prenom"),
    "residence": ("grande residence",),
    "hr_responsible": ("resp rh", "responsable rh"),
    "direction": ("dir c/r", "direction", "direction regionale"),
    "email": ("email", "mail", "adresse email", "adresse mail"),
}

MIN_SESSION_COLUMNS = {"module", "start_date", "end_date", "location"}


class PlanningParseError(ValueError):
    """Raised when an Excel planning file cannot be parsed."""


class PlanningExcelParser:
    """Parse Tunisie Telecom training planning workbooks."""

    def parse_workbook(self, filename: str, content: bytes) -> PlanningFileResult:
        if not filename.lower().endswith(".xlsx"):
            return PlanningFileResult(
                filename=filename,
                status="error",
                sheets=[],
                errors=[
                    "Unsupported planning file format. Please upload an Excel .xlsx file."
                ],
            )

        try:
            workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        except Exception as exc:
            return PlanningFileResult(
                filename=filename,
                status="error",
                sheets=[],
                errors=[
                    "The Excel planning file could not be read. Check that it is not "
                    f"protected or corrupted, then upload it again. Technical detail: {exc}"
                ],
            )

        all_sessions: dict[str, TrainingSession] = {}
        warnings: list[str] = []
        errors: list[str] = []

        for worksheet in workbook.worksheets:
            try:
                sheet_sessions, sheet_warnings = self._parse_sheet(filename, worksheet)
                warnings.extend(sheet_warnings)
                for session in sheet_sessions:
                    existing = all_sessions.get(session.session_key)
                    if existing is None:
                        all_sessions[session.session_key] = session
                        continue
                    existing.source_rows.extend(session.source_rows)
                    existing.participants.extend(session.participants)
            except PlanningParseError as exc:
                warnings.append(f"{worksheet.title}: {exc}")
            except Exception as exc:
                errors.append(
                    f"{worksheet.title}: this sheet could not be imported. "
                    f"Please check the planning columns and dates. Technical detail: {exc}"
                )

        sessions = list(all_sessions.values())
        for session in sessions:
            self._validate_session(session)

        status = "ok"
        if errors:
            status = "error"
        elif warnings or any(session.missing_fields for session in sessions) or any(
            participant.missing_fields
            for session in sessions
            for participant in session.participants
        ):
            status = "needs_review"

        return PlanningFileResult(
            filename=filename,
            status=status,
            sheets=workbook.sheetnames,
            sessions=sessions,
            warnings=warnings,
            errors=errors,
        )

    def _parse_sheet(
        self,
        filename: str,
        worksheet: Worksheet,
    ) -> tuple[list[TrainingSession], list[str]]:
        header_row, column_map = self._detect_header(worksheet)
        if header_row is None:
            raise PlanningParseError(
                "Planning headers were not found. Expected columns include Module, "
                "Date Debut, Date Fin, Lieu de formation, Matricule, and Nom & Prenom."
            )

        sessions: dict[str, TrainingSession] = {}
        warnings: list[str] = []
        max_column = max(column_map.values(), default=0)
        rows = worksheet.iter_rows(
            min_row=header_row + 1,
            max_col=max_column,
            values_only=True,
        )

        for row_index, row in enumerate(rows, start=header_row + 1):
            row_data = self._row_values_from_tuple(row, column_map)
            if not self._has_session_data(row_data):
                continue

            session = self._session_from_row(
                filename=filename,
                worksheet=worksheet,
                row_index=row_index,
                row_data=row_data,
            )
            existing = sessions.get(session.session_key)
            if existing is None:
                sessions[session.session_key] = session
                existing = session
            else:
                existing.source_rows.append(row_index)

            participant = self._participant_from_row(row_index, row_data)
            if participant is not None:
                existing.participants.append(participant)

        if not sessions:
            warnings.append(
                f"{worksheet.title}: headers were found on row {header_row}, "
                "but no training sessions were imported. Check that session rows are filled."
            )

        return list(sessions.values()), warnings

    def _detect_header(self, worksheet: Worksheet) -> tuple[int | None, dict[str, int]]:
        best_row: int | None = None
        best_map: dict[str, int] = {}
        best_score = 0
        scan_limit = min(25, worksheet.max_row or 25)

        for row_index in range(1, scan_limit + 1):
            current_map: dict[str, int] = {}
            for cell in worksheet[row_index]:
                normalized = _normalize_header(cell.value)
                if not normalized:
                    continue
                for field_name, aliases in HEADER_ALIASES.items():
                    if field_name in current_map:
                        continue
                    if any(alias in normalized for alias in aliases):
                        current_map[field_name] = cell.column
                        break
            score = len(set(current_map) & (MIN_SESSION_COLUMNS | {"code_session", "module_code"}))
            if score > best_score:
                best_row = row_index
                best_map = current_map
                best_score = score

        if best_score < 2:
            return None, {}
        return best_row, best_map

    def _row_values_from_tuple(
        self,
        row: tuple[Any, ...],
        column_map: dict[str, int],
    ) -> dict[str, str]:
        values: dict[str, str] = {}
        for field_name, column_index in column_map.items():
            values[field_name] = _format_cell_value(
                row[column_index - 1] if column_index - 1 < len(row) else None
            )
        return values

    def _has_session_data(self, row_data: dict[str, str]) -> bool:
        return any(
            row_data.get(field, "").strip()
            for field in (
                "code_session",
                "module",
                "module_code",
                "project",
                "start_date",
                "full_name",
                "matricule",
            )
        )

    def _session_from_row(
        self,
        filename: str,
        worksheet: Worksheet,
        row_index: int,
        row_data: dict[str, str],
    ) -> TrainingSession:
        code_session = row_data.get("code_session", "")
        module = row_data.get("module", "")
        start_date = row_data.get("start_date", "")
        end_date = row_data.get("end_date", "")
        location = row_data.get("location", "")
        session_key = self._session_key(code_session, module, start_date, end_date, location)

        return TrainingSession(
            session_key=session_key,
            code_session=code_session,
            lms_session_number=row_data.get("lms_session_number", ""),
            malek_number=row_data.get("malek_number", ""),
            status=_normalize_status(row_data.get("status", "")),
            axis=row_data.get("axis", ""),
            domain=row_data.get("domain", ""),
            project=row_data.get("project", ""),
            training_type=row_data.get("training_type", ""),
            training_mode=row_data.get("training_mode", ""),
            certification_nature=row_data.get("certification_nature", ""),
            module_code=row_data.get("module_code", ""),
            module=module,
            cabinet=row_data.get("cabinet", ""),
            trainer=row_data.get("trainer", ""),
            selected_trainer=row_data.get("selected_trainer", ""),
            year=row_data.get("year", ""),
            month=row_data.get("month", ""),
            week=row_data.get("week", ""),
            duration_days=row_data.get("duration_days", ""),
            start_date=start_date,
            end_date=end_date,
            schedule=row_data.get("schedule", ""),
            hours_per_day=row_data.get("hours_per_day", ""),
            total_hours=row_data.get("total_hours", ""),
            location=location,
            accommodation_location=row_data.get("accommodation_location", ""),
            responsible_engagement=row_data.get("responsible_engagement", ""),
            candidate_count=row_data.get("candidate_count", ""),
            source_file=filename,
            source_sheet=worksheet.title,
            source_rows=[row_index],
        )

    def _participant_from_row(
        self,
        row_index: int,
        row_data: dict[str, str],
    ) -> PlanningParticipant | None:
        matricule = row_data.get("matricule", "")
        full_name = row_data.get("full_name", "")
        email = row_data.get("email", "")
        if not any((matricule, full_name, email)):
            return None

        missing_fields: list[str] = []
        if not email:
            missing_fields.append("email")
        if not full_name:
            missing_fields.append("full_name")

        return PlanningParticipant(
            matricule=matricule,
            full_name=full_name,
            email=email,
            residence=row_data.get("residence", ""),
            direction=row_data.get("direction", ""),
            hr_responsible=row_data.get("hr_responsible", ""),
            source_row=row_index,
            missing_fields=missing_fields,
        )

    def _validate_session(self, session: TrainingSession) -> None:
        missing = []
        if not session.module:
            missing.append("module")
        if not session.start_date:
            missing.append("start_date")
        if not session.end_date:
            missing.append("end_date")
        if not session.location:
            missing.append("location")
        session.missing_fields = missing

    def _session_key(
        self,
        code_session: str,
        module: str,
        start_date: str,
        end_date: str,
        location: str,
    ) -> str:
        raw = "|".join(
            value.strip().lower()
            for value in (code_session, module, start_date, end_date, location)
            if value.strip()
        )
        if not raw:
            raw = "unknown-session"
        return hashlib.sha1(raw.encode("utf-8")).hexdigest()[:16]


def _normalize_header(value: Any) -> str:
    text = _strip_accents(_format_cell_value(value)).lower()
    text = text.replace("°", " ").replace("º", " ")
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def _strip_accents(value: str) -> str:
    return "".join(
        char
        for char in unicodedata.normalize("NFKD", value)
        if not unicodedata.combining(char)
    )


def _format_cell_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.strftime("%H:%M")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _normalize_status(value: str) -> str:
    normalized = _strip_accents(value).strip().lower()
    if not normalized:
        return "UNKNOWN"
    status_map = {
        "a": "CANCELLED",
        "annulee": "CANCELLED",
        "annule": "CANCELLED",
        "r": "DONE",
        "realisee": "DONE",
        "realise": "DONE",
        "rep": "POSTPONED",
        "reportee": "POSTPONED",
        "reporte": "POSTPONED",
        "p": "PLANNED_SENT",
        "planifiee": "PLANNED",
        "planifie": "PLANNED",
        "pc": "CONFIRMED",
    }
    return status_map.get(normalized, normalized.upper())


def flatten_sessions(files: Iterable[PlanningFileResult]) -> list[TrainingSession]:
    return [session for file_result in files for session in file_result.sessions]
