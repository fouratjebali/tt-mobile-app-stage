from __future__ import annotations

import csv
import re
import unicodedata
from dataclasses import asdict, dataclass, field
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


CONTACT_HEADER_ALIASES: dict[str, tuple[str, ...]] = {
    "matricule": ("matricule", "matricules", "identifiant", "id agent"),
    "full_name": ("nom & prenom", "nom et prenom", "nom prenom", "nom", "prenom"),
    "email": ("email", "mail", "adresse email", "adresse mail", "e-mail"),
    "direction": ("direction", "direction regionale", "dir c/r", "structure"),
    "hr_responsible": ("resp rh", "responsable rh", "rh"),
}


@dataclass(slots=True)
class EmployeeContact:
    matricule: str = ""
    full_name: str = ""
    email: str = ""
    direction: str = ""
    hr_responsible: str = ""
    source_file: str = ""
    source_row: int = 0

    @property
    def normalized_name(self) -> str:
        return normalize_name(self.full_name)

    def to_dict(self) -> dict[str, Any]:
        data = asdict(self)
        data["normalized_name"] = self.normalized_name
        return data


@dataclass(slots=True)
class ContactImportResult:
    filename: str
    status: str
    contacts: list[EmployeeContact] = field(default_factory=list)
    imported_count: int = 0
    skipped_count: int = 0
    warnings: list[str] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)

    def to_dict(self) -> dict[str, Any]:
        data = asdict(self)
        data["contacts"] = [contact.to_dict() for contact in self.contacts]
        return data


class ContactDirectoryParser:
    def parse_file(self, filename: str, content: bytes) -> ContactImportResult:
        suffix = Path(filename).suffix.lower()
        if suffix == ".xlsx":
            return self._parse_xlsx(filename, content)
        if suffix == ".csv":
            return self._parse_csv(filename, content)
        return ContactImportResult(
            filename=filename,
            status="error",
            errors=[
                "Unsupported contact file format. Please upload a .xlsx or .csv directory."
            ],
        )

    def _parse_xlsx(self, filename: str, content: bytes) -> ContactImportResult:
        try:
            workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        except Exception as exc:
            return ContactImportResult(
                filename=filename,
                status="error",
                errors=[
                    "The contact directory could not be read. Check that it is not "
                    f"protected or corrupted, then upload it again. Technical detail: {exc}"
                ],
            )

        contacts: list[EmployeeContact] = []
        warnings: list[str] = []
        skipped = 0
        for worksheet in workbook.worksheets:
            header_row, column_map = self._detect_xlsx_header(worksheet)
            if header_row is None:
                warnings.append(
                    f"{worksheet.title}: contact headers were not found. "
                    "Expected columns include Email plus Nom & Prenom or Matricule."
                )
                continue
            max_column = max(column_map.values(), default=0)
            for row_index, row in enumerate(
                worksheet.iter_rows(
                    min_row=header_row + 1,
                    max_col=max_column,
                    values_only=True,
                ),
                start=header_row + 1,
            ):
                row_data = {
                    field_name: _cell_text(row[column_index - 1] if column_index - 1 < len(row) else None)
                    for field_name, column_index in column_map.items()
                }
                contact = self._contact_from_row(filename, row_index, row_data)
                if contact is None:
                    skipped += 1
                    continue
                contacts.append(contact)

        return _result(filename, contacts, skipped, warnings)

    def _parse_csv(self, filename: str, content: bytes) -> ContactImportResult:
        text = content.decode("utf-8-sig", errors="replace")
        sample = text[:2048]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;")
        except csv.Error:
            dialect = csv.excel
        reader = csv.reader(StringIO(text), dialect)
        rows = list(reader)
        if not rows:
            return ContactImportResult(
                filename=filename,
                status="error",
                errors=["The contact directory is empty. Choose a file that contains employee rows."],
            )

        header_index, column_map = self._detect_csv_header(rows)
        if header_index is None:
            return ContactImportResult(
                filename=filename,
                status="error",
                errors=[
                    "Contact headers were not found. Expected columns include Email "
                    "plus Nom & Prenom or Matricule."
                ],
            )

        contacts: list[EmployeeContact] = []
        skipped = 0
        for index, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
            row_data = {
                field_name: _cell_text(row[column_index] if column_index < len(row) else None)
                for field_name, column_index in column_map.items()
            }
            contact = self._contact_from_row(filename, index, row_data)
            if contact is None:
                skipped += 1
                continue
            contacts.append(contact)

        return _result(filename, contacts, skipped, [])

    def _detect_xlsx_header(self, worksheet: Any) -> tuple[int | None, dict[str, int]]:
        best_row: int | None = None
        best_map: dict[str, int] = {}
        best_score = 0
        for row_index in range(1, min(20, worksheet.max_row or 20) + 1):
            current_map: dict[str, int] = {}
            for cell in worksheet[row_index]:
                normalized = normalize_header(cell.value)
                self._map_header_cell(normalized, cell.column, current_map)
            score = self._header_score(current_map)
            if score > best_score:
                best_row = row_index
                best_map = current_map
                best_score = score
        if best_score < 2 or "email" not in best_map:
            return None, {}
        return best_row, best_map

    def _detect_csv_header(self, rows: list[list[str]]) -> tuple[int | None, dict[str, int]]:
        best_index: int | None = None
        best_map: dict[str, int] = {}
        best_score = 0
        for index, row in enumerate(rows[:20]):
            current_map: dict[str, int] = {}
            for column_index, value in enumerate(row):
                normalized = normalize_header(value)
                self._map_header_cell(normalized, column_index, current_map)
            score = self._header_score(current_map)
            if score > best_score:
                best_index = index
                best_map = current_map
                best_score = score
        if best_score < 2 or "email" not in best_map:
            return None, {}
        return best_index, best_map

    def _map_header_cell(
        self,
        normalized: str,
        column_index: int,
        current_map: dict[str, int],
    ) -> None:
        if not normalized:
            return
        for field_name, aliases in CONTACT_HEADER_ALIASES.items():
            if field_name in current_map:
                continue
            if any(alias in normalized for alias in aliases):
                current_map[field_name] = column_index
                break

    def _header_score(self, column_map: dict[str, int]) -> int:
        score = len(column_map)
        if "email" in column_map:
            score += 2
        if "matricule" in column_map or "full_name" in column_map:
            score += 2
        return score

    def _contact_from_row(
        self,
        filename: str,
        row_index: int,
        row_data: dict[str, str],
    ) -> EmployeeContact | None:
        email = row_data.get("email", "").strip().lower()
        matricule = row_data.get("matricule", "").strip()
        full_name = clean_name(row_data.get("full_name", ""))
        if not email or "@" not in email:
            return None
        if not matricule and not full_name:
            return None
        return EmployeeContact(
            matricule=matricule,
            full_name=full_name,
            email=email,
            direction=row_data.get("direction", ""),
            hr_responsible=row_data.get("hr_responsible", ""),
            source_file=filename,
            source_row=row_index,
        )


def _result(
    filename: str,
    contacts: list[EmployeeContact],
    skipped: int,
    warnings: list[str],
) -> ContactImportResult:
    status = "ok" if contacts else "error"
    errors = (
        []
        if contacts
        else [
            "No valid contacts were found. Each row needs an email and either "
            "a matricule or a name."
        ]
    )
    return ContactImportResult(
        filename=filename,
        status=status,
        contacts=contacts,
        imported_count=len(contacts),
        skipped_count=skipped,
        warnings=warnings,
        errors=errors,
    )


def clean_name(value: str) -> str:
    return re.sub(r"\s+", " ", value).strip()


def normalize_name(value: str) -> str:
    text = strip_accents(clean_name(value)).lower()
    text = re.sub(r"[^a-z0-9 ]+", " ", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def normalize_header(value: Any) -> str:
    text = strip_accents(_cell_text(value)).lower()
    text = text.replace("°", " ").replace("º", " ")
    text = re.sub(r"[^a-z0-9/& ]+", " ", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def strip_accents(value: str) -> str:
    return "".join(
        char
        for char in unicodedata.normalize("NFKD", value)
        if not unicodedata.combining(char)
    )


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()
